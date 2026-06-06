import psycopg2
from psycopg2.pool import ThreadedConnectionPool
from datetime import datetime, timedelta, date
import calendar
import pytz
import openpyxl
from openpyxl.styles import Font
import os
import time

TASHKENT = pytz.timezone('Asia/Tashkent')
SUPER_ADMIN_KOD = os.environ.get("SUPER_ADMIN_KOD", "0001")

# Railway asosiy baza, Neon zaxira baza
RAILWAY_URL = "postgresql://postgres:RdcrgixOGANtWvspNqPdFVPhyUkBmjeS@kodama.proxy.rlwy.net:59039/railway"
NEON_URL = "postgresql://neondb_owner:npg_mlJ7sC9LhqTB@ep-snowy-cake-apn5lmrf-pooler.c-7.us-east-1.aws.neon.tech/neondb?sslmode=require"
DATABASE_URL = os.environ.get("DATABASE_URL", NEON_URL)

_pool = None
_komp_kesh = {}   # {komp_id: (vaqt, data)}
_KESH_TTL = 30    # 30 soniya
_railway_online = False
_last_railway_check = 0
_RAILWAY_CHECK_INTERVAL = 60  # har 60 soniyada Railway ni tekshir

def _check_railway():
    """Railway bazasi qayta ishlayaptimi tekshirish"""
    global _railway_online, _last_railway_check, DATABASE_URL, _pool
    now = time.time()
    if now - _last_railway_check < _RAILWAY_CHECK_INTERVAL:
        return
    _last_railway_check = now
    try:
        test = psycopg2.connect(RAILWAY_URL, connect_timeout=5)
        test.close()
        if not _railway_online:
            print("✅ RAILWAY BAZA QAYTA ISHLADI! Ulanmoqda...")
            _railway_online = True
            DATABASE_URL = RAILWAY_URL
            if _pool and not _pool.closed:
                _pool.closeall()
            _pool = None
    except:
        _railway_online = False

def _get_pool():
    global _pool
    _check_railway()
    if _pool is None or _pool.closed:
        _pool = ThreadedConnectionPool(1, 5, DATABASE_URL)
    return _pool

class _PooledConn:
    def __init__(self, conn):
        self._conn = conn
    def cursor(self):
        return self._conn.cursor()
    def commit(self):
        return self._conn.commit()
    def rollback(self):
        return self._conn.rollback()
    def close(self):
        _get_pool().putconn(self._conn)

def connect():
    return _PooledConn(_get_pool().getconn())

def hozir():
    return datetime.now(TASHKENT)

def soat_format(v):
    v = float(v or 0)
    s = int(v); d = int((v - s) * 60)
    return f"{s} soat {d} daqiqa"

def kechikish_format(d):
    d = int(d or 0)
    return f"{d//60} soat {d%60} daqiqa" if d >= 60 else f"{d} daqiqa"

def create_tables():
    import time as _t
    for attempt in range(3):
        try:
            _create_tables_inner()
            return
        except Exception as e:
            if 'deadlock' in str(e).lower() and attempt < 2:
                _t.sleep(2)
            else:
                raise

def _create_tables_inner():
    conn = connect(); cur = conn.cursor()

    cur.execute('''CREATE TABLE IF NOT EXISTS super_adminlar (
        id SERIAL PRIMARY KEY, telefon TEXT NOT NULL,
        telegram_id BIGINT, ism TEXT DEFAULT 'Super Admin', kod TEXT DEFAULT '0001'
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS kompaniyalar (
        id SERIAL PRIMARY KEY, nomi TEXT NOT NULL,
        admin_telefon TEXT, admin_id BIGINT, admin_kod TEXT DEFAULT '1234',
        gps_lat REAL DEFAULT 41.299496, gps_lon REAL DEFAULT 69.240073,
        gps_radius INTEGER DEFAULT 200, holat TEXT DEFAULT 'faol', yaratilgan TEXT,
        gps_aktiv BOOLEAN DEFAULT TRUE, selfie_aktiv BOOLEAN DEFAULT TRUE,
        face_id_aktiv BOOLEAN DEFAULT FALSE, hikvision_aktiv BOOLEAN DEFAULT FALSE,
        live_gps_aktiv BOOLEAN DEFAULT FALSE, live_gps_tekshiruv BOOLEAN DEFAULT FALSE
    )''')
    cur.execute("ALTER TABLE kompaniyalar ADD COLUMN IF NOT EXISTS admin_kod TEXT DEFAULT '1234'")
    cur.execute("ALTER TABLE kompaniyalar ADD COLUMN IF NOT EXISTS live_gps_aktiv BOOLEAN DEFAULT FALSE")
    cur.execute("ALTER TABLE kompaniyalar ADD COLUMN IF NOT EXISTS live_gps_tekshiruv BOOLEAN DEFAULT FALSE")
    cur.execute("ALTER TABLE kompaniyalar ADD COLUMN IF NOT EXISTS wifi_aktiv BOOLEAN DEFAULT FALSE")
    cur.execute("ALTER TABLE kompaniyalar ADD COLUMN IF NOT EXISTS wifi_ssid TEXT DEFAULT ''")

    # Migration: rename wifi_ip to wifi_mac if wifi_ip exists but wifi_mac doesn't
    try:
        cur.execute("SELECT column_name FROM information_schema.columns WHERE table_name='kompaniyalar' AND column_name='wifi_ip'")
        if cur.fetchone():
            cur.execute("ALTER TABLE kompaniyalar RENAME COLUMN wifi_ip TO wifi_mac")
    except:
        pass

    # If migration didn't work, just add wifi_mac column
    cur.execute("ALTER TABLE kompaniyalar ADD COLUMN IF NOT EXISTS wifi_mac TEXT DEFAULT ''")
    cur.execute("ALTER TABLE kompaniyalar ADD COLUMN IF NOT EXISTS audio_aktiv BOOLEAN DEFAULT FALSE")
    cur.execute("ALTER TABLE kompaniyalar ADD COLUMN IF NOT EXISTS matn_aktiv BOOLEAN DEFAULT FALSE")

    cur.execute('''CREATE TABLE IF NOT EXISTS xodimlar (
        id SERIAL PRIMARY KEY, ism TEXT NOT NULL, telefon TEXT, kod TEXT,
        lavozim TEXT, oylik REAL DEFAULT 0,
        ish_boshlanish TEXT DEFAULT '09:00', ish_tugash TEXT DEFAULT '18:00',
        ishga_kirgan TEXT, telegram_id BIGINT UNIQUE,
        kompaniya_id INTEGER REFERENCES kompaniyalar(id),
        rol TEXT DEFAULT 'xodim', holat TEXT DEFAULT 'faol', tugilgan_kun TEXT
    )''')
    cur.execute("ALTER TABLE xodimlar ADD COLUMN IF NOT EXISTS tugilgan_kun TEXT")
    # Xodim uchun alohida funksiya sozlamalari (NULL = kompaniya sozlamasiga bo'ysunadi)
    for col in ['xod_gps','xod_selfie','xod_audio','xod_matn','xod_face_id','xod_hikvision','xod_live_gps','xod_wifi']:
        cur.execute(f"ALTER TABLE xodimlar ADD COLUMN IF NOT EXISTS {col} BOOLEAN DEFAULT NULL")

    cur.execute('''CREATE TABLE IF NOT EXISTS xodim_kpi (
        id SERIAL PRIMARY KEY,
        xodim_id INTEGER REFERENCES xodimlar(id) ON DELETE CASCADE,
        komp_id INTEGER REFERENCES kompaniyalar(id),
        topshiriq_jarima DECIMAL DEFAULT 0,
        sababsiz_qolish_soat INTEGER DEFAULT 0,
        sababsiz_qolish_jarima DECIMAL DEFAULT 0,
        sababli_qolish_soat INTEGER DEFAULT 0,
        sababli_qolish_jarima DECIMAL DEFAULT 0,
        kech_kelish_daqiqa INTEGER DEFAULT 0,
        kech_kelish_jarima DECIMAL DEFAULT 0,
        erta_ketish_daqiqa INTEGER DEFAULT 0,
        erta_ketish_jarima DECIMAL DEFAULT 0,
        joy_tark_soat INTEGER DEFAULT 0,
        joy_tark_jarima DECIMAL DEFAULT 0,
        alo_mukofot DECIMAL DEFAULT 0,
        qoshimcha_jarima DECIMAL DEFAULT 0
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS davomat (
        id SERIAL PRIMARY KEY, xodim_id INTEGER REFERENCES xodimlar(id),
        kompaniya_id INTEGER REFERENCES kompaniyalar(id),
        sana TEXT, keldi TEXT, ketdi TEXT, ish_soat REAL DEFAULT 0,
        kechikish INTEGER DEFAULT 0, holat TEXT DEFAULT 'normal',
        izoh TEXT, keldi_rasm TEXT, ketdi_rasm TEXT,
        kiritdi TEXT DEFAULT 'xodim', kiritdi_id BIGINT
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS sababli_sorovlar (
        id SERIAL PRIMARY KEY, xodim_id INTEGER REFERENCES xodimlar(id),
        kompaniya_id INTEGER REFERENCES kompaniyalar(id),
        sana TEXT, sabab TEXT, holat TEXT DEFAULT 'kutilmoqda'
    )''')
    cur.execute("ALTER TABLE sababli_sorovlar ADD COLUMN IF NOT EXISTS kompaniya_id INTEGER REFERENCES kompaniyalar(id)")

    cur.execute('''CREATE TABLE IF NOT EXISTS live_lokatsiyalar (
        id SERIAL PRIMARY KEY, xodim_id INTEGER REFERENCES xodimlar(id),
        kompaniya_id INTEGER REFERENCES kompaniyalar(id),
        lat REAL, lon REAL, vaqt TEXT, faol BOOLEAN DEFAULT TRUE
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS sozlamalar (
        kalit TEXT PRIMARY KEY, qiymat TEXT
    )''')
    cur.execute("INSERT INTO sozlamalar(kalit,qiymat) VALUES('murojaat_raqam','919712222') ON CONFLICT(kalit) DO NOTHING")

    cur.execute('''CREATE TABLE IF NOT EXISTS audit_log (
        id SERIAL PRIMARY KEY, kompaniya_id INTEGER REFERENCES kompaniyalar(id),
        xodim_id INTEGER REFERENCES xodimlar(id),
        amal TEXT, tafsilot TEXT, vaqt TEXT, rasm_id TEXT, video_id TEXT,
        user_id BIGINT, user_ism TEXT
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS wifi_macs (
        id SERIAL PRIMARY KEY, kompaniya_id INTEGER REFERENCES kompaniyalar(id),
        mac_address TEXT NOT NULL, nomi TEXT DEFAULT ''
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS xabarlar (
        id SERIAL PRIMARY KEY,
        from_user_id BIGINT,
        from_user_name TEXT,
        from_role TEXT,
        from_org_id INTEGER,
        to_user_id BIGINT,
        to_user_name TEXT,
        to_role TEXT,
        to_org_id INTEGER,
        subject TEXT,
        xabar TEXT,
        rasm_id TEXT,
        dokument_id TEXT,
        holat TEXT DEFAULT 'o''qilmadi',
        yaratilgan TEXT,
        o_qilgan TEXT,
        org_id INTEGER NOT NULL
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS xabar_qabul_qiluvchilar (
        id SERIAL PRIMARY KEY,
        xabar_id INTEGER REFERENCES xabarlar(id),
        qabul_qiluvchi_id BIGINT,
        qabul_qiluvchi_nomi TEXT,
        qabul_qiluvchi_role TEXT,
        holat TEXT DEFAULT 'o''qilmadi',
        o_qilgan TEXT
    )''')

    # ========== KIRM (INCOME) ==========
    cur.execute('''CREATE TABLE IF NOT EXISTS kirm (
        id SERIAL PRIMARY KEY,
        xodim_id INTEGER NOT NULL,
        komp_id INTEGER NOT NULL,
        turi TEXT NOT NULL,
        summa DECIMAL(12,2) NOT NULL,
        izoh TEXT,
        sana TEXT NOT NULL,
        vaqt TEXT NOT NULL,
        yaratilgan TEXT,
        UPDATE_BY_ROLE TEXT,
        UPDATE_BY_NAME TEXT,
        holat TEXT DEFAULT 'qabul_qilindi'
    )''')

    # ========== CHIQIM (EXPENSE/PENALTY) ==========
    cur.execute('''CREATE TABLE IF NOT EXISTS chiqim (
        id SERIAL PRIMARY KEY,
        xodim_id INTEGER NOT NULL,
        komp_id INTEGER NOT NULL,
        turi TEXT NOT NULL,
        summa DECIMAL(12,2) NOT NULL,
        izoh TEXT,
        sana TEXT NOT NULL,
        vaqt TEXT NOT NULL,
        yaratilgan TEXT,
        UPDATE_BY_ROLE TEXT,
        UPDATE_BY_NAME TEXT,
        holat TEXT DEFAULT 'qabul_qilindi'
    )''')

    conn.commit(); cur.close(); conn.close()

# ========== SOZLAMALAR ==========

def get_murojaat_raqam():
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT qiymat FROM sozlamalar WHERE kalit='murojaat_raqam'")
    r = cur.fetchone(); cur.close(); conn.close()
    return r[0] if r else '919712222'

def set_murojaat_raqam(raqam):
    conn = connect(); cur = conn.cursor()
    cur.execute("INSERT INTO sozlamalar(kalit,qiymat) VALUES('murojaat_raqam',%s) ON CONFLICT(kalit) DO UPDATE SET qiymat=%s", (raqam, raqam))
    conn.commit(); cur.close(); conn.close()

# ========== SUPER ADMIN ==========

def super_admin_tekshir(telefon):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id,ism,telegram_id FROM super_adminlar WHERE telefon LIKE %s", (f"%{telefon[-9:]}%",))
    r = cur.fetchone(); cur.close(); conn.close(); return r

def super_admin_kod_tekshir(kod):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT kod FROM super_adminlar LIMIT 1")
    r = cur.fetchone(); cur.close(); conn.close()
    db_kod = r[0] if r else SUPER_ADMIN_KOD
    return kod == db_kod

def super_admin_telegram_saqlash(telefon, telegram_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id FROM super_adminlar WHERE telefon LIKE %s", (f"%{telefon[-9:]}%",))
    if cur.fetchone():
        cur.execute("UPDATE super_adminlar SET telegram_id=%s WHERE telefon LIKE %s", (telegram_id, f"%{telefon[-9:]}%"))
    else:
        cur.execute("INSERT INTO super_adminlar(telefon,telegram_id) VALUES(%s,%s)", (telefon, telegram_id))
    conn.commit(); cur.close(); conn.close()

def super_admin_id_tekshir(telegram_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id FROM super_adminlar WHERE telegram_id=%s", (telegram_id,))
    r = cur.fetchone(); cur.close(); conn.close(); return r is not None

def super_admin_kod_ozgartir(yangi_kod):
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE super_adminlar SET kod=%s", (yangi_kod,))
    conn.commit(); cur.close(); conn.close()

def super_admin_telefon_ozgartir(telegram_id, tel):
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE super_adminlar SET telefon=%s WHERE telegram_id=%s", (tel, telegram_id))
    conn.commit(); cur.close(); conn.close()

def super_admin_ism_ozgartir(telegram_id, ism):
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE super_adminlar SET ism=%s WHERE telegram_id=%s", (ism, telegram_id))
    conn.commit(); cur.close(); conn.close()

def barcha_super_adminlar():
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id,telefon,telegram_id,ism FROM super_adminlar ORDER BY id")
    r = cur.fetchall(); cur.close(); conn.close(); return r

def barcha_super_admin_idlar():
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT telegram_id FROM super_adminlar WHERE telegram_id IS NOT NULL")
    r = cur.fetchall(); cur.close(); conn.close(); return [x[0] for x in r]

def super_admin_qoshish(telefon, ism='Super Admin'):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id FROM super_adminlar WHERE telefon LIKE %s", (f"%{telefon[-9:]}%",))
    if cur.fetchone(): cur.close(); conn.close(); return False
    cur.execute("INSERT INTO super_adminlar(telefon,ism) VALUES(%s,%s)", (telefon, ism))
    conn.commit(); cur.close(); conn.close(); return True

def super_admin_ochirish(sa_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("DELETE FROM super_adminlar WHERE id=%s", (sa_id,))
    conn.commit(); cur.close(); conn.close()

def super_admin_olish(telegram_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id,telefon,ism FROM super_adminlar WHERE telegram_id=%s", (telegram_id,))
    r = cur.fetchone(); cur.close(); conn.close(); return r

# ========== KOMPANIYALAR ==========

def kompaniya_yaratish(nomi, admin_telefon, admin_kod='1234'):
    conn = connect(); cur = conn.cursor()
    sana = hozir().strftime("%Y-%m-%d")
    cur.execute("INSERT INTO kompaniyalar(nomi,admin_telefon,admin_kod,yaratilgan) VALUES(%s,%s,%s,%s) RETURNING id",
                (nomi, admin_telefon, admin_kod, sana))
    kid = cur.fetchone()[0]; conn.commit(); cur.close(); conn.close(); return kid

def kompaniya_ochirish(komp_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("DELETE FROM live_lokatsiyalar WHERE kompaniya_id=%s", (komp_id,))
    cur.execute("DELETE FROM sababli_sorovlar WHERE kompaniya_id=%s", (komp_id,))
    cur.execute("DELETE FROM davomat WHERE kompaniya_id=%s", (komp_id,))
    cur.execute("DELETE FROM xodimlar WHERE kompaniya_id=%s", (komp_id,))
    cur.execute("DELETE FROM kompaniyalar WHERE id=%s", (komp_id,))
    conn.commit(); cur.close(); conn.close()

def barcha_kompaniyalar():
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT id,nomi,admin_telefon,holat,yaratilgan,
                  gps_aktiv,selfie_aktiv,face_id_aktiv,hikvision_aktiv,admin_kod,
                  live_gps_aktiv,live_gps_tekshiruv FROM kompaniyalar ORDER BY id''')
    r = cur.fetchall(); cur.close(); conn.close(); return r

def kompaniya_olish(komp_id, keshdan=True):
    if keshdan and komp_id in _komp_kesh:
        ts, data = _komp_kesh[komp_id]
        if time.time() - ts < _KESH_TTL:
            return data
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT id,nomi,admin_telefon,admin_id,holat,
                  gps_lat,gps_lon,gps_radius,yaratilgan,
                  gps_aktiv,selfie_aktiv,face_id_aktiv,hikvision_aktiv,admin_kod,
                  live_gps_aktiv,live_gps_tekshiruv,wifi_aktiv,wifi_ssid,wifi_mac,
                  COALESCE(audio_aktiv,FALSE),COALESCE(matn_aktiv,FALSE)
                  FROM kompaniyalar WHERE id=%s''', (komp_id,))
    r = cur.fetchone(); cur.close(); conn.close()
    if r: _komp_kesh[komp_id] = (time.time(), r)
    return r

def kompaniya_kesh_tozala(komp_id=None):
    if komp_id: _komp_kesh.pop(komp_id, None)
    else: _komp_kesh.clear()

def kompaniya_holat_ozgartir(komp_id, holat):
    try:
        conn = connect(); cur = conn.cursor()
        cur.execute("UPDATE kompaniyalar SET holat=%s WHERE id=%s", (holat, komp_id))
        rows_updated = cur.rowcount
        conn.commit(); cur.close(); conn.close()
        if rows_updated == 0:
            raise Exception(f"No company found with id {komp_id}")
        return True
    except Exception as e:
        if 'conn' in locals():
            try:
                conn.rollback()
            except:
                pass
            cur.close()
            conn.close()
        raise Exception(f"Failed to update company status: {str(e)}")

def kompaniya_funksiya_ozgartir(komp_id, funksiya, qiymat):
    conn = connect(); cur = conn.cursor()
    cur.execute(f"UPDATE kompaniyalar SET {funksiya}=%s WHERE id=%s", (qiymat, komp_id))
    conn.commit(); cur.close(); conn.close()
    kompaniya_kesh_tozala(komp_id)

def kompaniya_tahrirlash(komp_id, maydon, qiymat):
    conn = connect(); cur = conn.cursor()
    cur.execute(f"UPDATE kompaniyalar SET {maydon}=%s WHERE id=%s", (qiymat, komp_id))
    conn.commit(); cur.close(); conn.close()

def admin_telefon_orqali_kompaniya(telefon):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id,nomi,holat,admin_kod FROM kompaniyalar WHERE admin_telefon LIKE %s", (f"%{telefon[-9:]}%",))
    r = cur.fetchone(); cur.close(); conn.close(); return r

def admin_id_saqlash(komp_id, admin_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE kompaniyalar SET admin_id=%s WHERE id=%s", (admin_id, komp_id))
    conn.commit(); cur.close(); conn.close()

def get_gps(komp_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT gps_lat,gps_lon,gps_radius FROM kompaniyalar WHERE id=%s", (komp_id,))
    r = cur.fetchone(); cur.close(); conn.close()
    return r if r else (41.299496, 69.240073, 200)

def get_wifi(komp_id):
    """WiFi sozlamalarini olish (SSID va MAC)"""
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT wifi_aktiv,wifi_ssid,wifi_mac FROM kompaniyalar WHERE id=%s", (komp_id,))
    r = cur.fetchone(); cur.close(); conn.close()
    return r if r else (False, "", "")

def get_wifi_mac(komp_id):
    """WiFi MAC manzilini olish"""
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT wifi_aktiv,wifi_mac FROM kompaniyalar WHERE id=%s", (komp_id,))
    r = cur.fetchone(); cur.close(); conn.close()
    return r if r else (False, "")

def wifi_sozla(komp_id, aktiv, ssid, mac=""):
    """WiFi sozlamalarini o'rnatish (SSID va MAC)"""
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE kompaniyalar SET wifi_aktiv=%s,wifi_ssid=%s,wifi_mac=%s WHERE id=%s",
                (aktiv, ssid, mac, komp_id))
    conn.commit(); cur.close(); conn.close()

def wifi_mac_qosh(komp_id, mac_address, nomi=""):
    """WiFi MAC manzilini qo'shish"""
    conn = connect(); cur = conn.cursor()
    cur.execute("INSERT INTO wifi_macs(kompaniya_id, mac_address, nomi) VALUES(%s, %s, %s)",
                (komp_id, mac_address.upper(), nomi))
    conn.commit(); cur.close(); conn.close()

def wifi_mac_olish(komp_id):
    """Kompaniyaning barcha WiFi MAC manzillarini olish"""
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id, mac_address, nomi FROM wifi_macs WHERE kompaniya_id=%s ORDER BY id", (komp_id,))
    r = cur.fetchall(); cur.close(); conn.close()
    return r if r else []

def wifi_mac_tekshir(komp_id, mac_address):
    """MAC manzil kompaniya ro'yxatida bor-yo'qligini tekshirish"""
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id FROM wifi_macs WHERE kompaniya_id=%s AND UPPER(mac_address)=%s",
                (komp_id, mac_address.upper()))
    r = cur.fetchone(); cur.close(); conn.close()
    return r is not None

def wifi_mac_ochir(mac_id):
    """WiFi MAC manzilini o'chirish"""
    conn = connect(); cur = conn.cursor()
    cur.execute("DELETE FROM wifi_macs WHERE id=%s", (mac_id,))
    conn.commit(); cur.close(); conn.close()

def wifi_mac_tahrir(mac_id, nomi):
    """WiFi MAC nomini tahrirlash"""
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE wifi_macs SET nomi=%s WHERE id=%s", (nomi, mac_id))
    conn.commit(); cur.close(); conn.close()

def komp_bugun_rasmlar(komp_id):
    """Kompaniya xodimlarining bugungi barcha rasmlari - to'liq ma'lumot bilan"""
    conn = connect(); cur = conn.cursor()
    sana = hozir().strftime("%Y-%m-%d")
    cur.execute('''SELECT x.id, x.ism, x.lavozim, d.keldi, d.keldi_rasm, a.tafsilot
                   FROM davomat d
                   JOIN xodimlar x ON d.xodim_id=x.id
                   LEFT JOIN audit_log a ON a.xodim_id=x.id AND a.amal='KELDI' AND DATE(a.vaqt)=%s
                   WHERE d.sana=%s AND x.komp_id=%s AND d.keldi_rasm IS NOT NULL
                   ORDER BY d.keldi DESC''', (sana, sana, komp_id))
    r = cur.fetchall(); cur.close(); conn.close(); return r

def komp_bugun_videolar(komp_id):
    """Kompaniya xodimlarining bugungi barcha videolari - to'liq ma'lumot bilan"""
    conn = connect(); cur = conn.cursor()
    sana = hozir().strftime("%Y-%m-%d")
    cur.execute('''SELECT x.id, x.ism, x.lavozim, d.ketdi, d.ketdi_rasm, a.tafsilot
                   FROM davomat d
                   JOIN xodimlar x ON d.xodim_id=x.id
                   LEFT JOIN audit_log a ON a.xodim_id=x.id AND a.amal='KETDI' AND DATE(a.vaqt)=%s
                   WHERE d.sana=%s AND x.komp_id=%s AND d.ketdi_rasm IS NOT NULL
                   ORDER BY d.ketdi DESC''', (sana, sana, komp_id))
    r = cur.fetchall(); cur.close(); conn.close(); return r

def barcha_komp_bugun_rasmlar():
    """Barcha kompaniyalarning bugungi rasmlari - super admin uchun - to'liq ma'lumot"""
    conn = connect(); cur = conn.cursor()
    sana = hozir().strftime("%Y-%m-%d")
    cur.execute('''SELECT k.nomi, x.ism, x.lavozim, d.keldi, d.keldi_rasm, a.tafsilot
                   FROM davomat d
                   JOIN xodimlar x ON d.xodim_id=x.id
                   JOIN kompaniyalar k ON x.komp_id=k.id
                   LEFT JOIN audit_log a ON a.xodim_id=x.id AND a.amal='KELDI' AND DATE(a.vaqt)=%s
                   WHERE d.sana=%s AND d.keldi_rasm IS NOT NULL
                   ORDER BY d.keldi DESC''', (sana, sana))
    r = cur.fetchall(); cur.close(); conn.close(); return r

# ========== XODIMLAR ==========

def xodim_qoshish(ism, telefon, lavozim, oylik, ish_bosh, ish_tug, komp_id, rol, kod):
    try:
        conn = connect(); cur = conn.cursor()
        sana = hozir().strftime("%Y-%m-%d")
        cur.execute('''INSERT INTO xodimlar(ism,telefon,lavozim,oylik,ish_boshlanish,
                      ish_tugash,kompaniya_id,rol,kod,ishga_kirgan)
                      VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id''',
                    (ism,telefon,lavozim,oylik,ish_bosh,ish_tug,komp_id,rol,kod,sana))
        xid = cur.fetchone()
        if xid:
            xid = xid[0]
            conn.commit()
            cur.close()
            conn.close()
            return xid
        else:
            conn.rollback()
            cur.close()
            conn.close()
            raise Exception("Failed to insert employee: No ID returned")
    except Exception as e:
        if 'conn' in locals():
            try:
                conn.rollback()
            except:
                pass
            cur.close()
            conn.close()
        raise Exception(f"Employee add failed: {str(e)}")

def xodim_ochirish(xodim_id):
    conn = connect(); cur = conn.cursor()
    # Birinchi, barcha related records'ni o'chir (FK constraints)
    cur.execute("DELETE FROM audit_log WHERE xodim_id=%s", (xodim_id,))
    cur.execute("DELETE FROM live_lokatsiyalar WHERE xodim_id=%s", (xodim_id,))
    cur.execute("DELETE FROM sababli_sorovlar WHERE xodim_id=%s", (xodim_id,))
    cur.execute("DELETE FROM davomat WHERE xodim_id=%s", (xodim_id,))
    # Keyin xodimni o'chir
    cur.execute("DELETE FROM xodimlar WHERE id=%s", (xodim_id,))
    conn.commit(); cur.close(); conn.close()

def kompaniya_xodimlari(komp_id):
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT id,ism,lavozim,telefon,oylik,ish_boshlanish,
                  ish_tugash,rol,kod,holat FROM xodimlar
                  WHERE kompaniya_id=%s ORDER BY id''', (komp_id,))
    r = cur.fetchall(); cur.close(); conn.close(); return r

def xodim_olish(xodim_id):
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT id,ism,telefon,lavozim,oylik,ish_boshlanish,
                  ish_tugash,rol,kod,kompaniya_id,holat,tugilgan_kun
                  FROM xodimlar WHERE id=%s''', (xodim_id,))
    r = cur.fetchone(); cur.close(); conn.close(); return r

def xodim_tahrirlash(xodim_id, maydon, qiymat):
    conn = connect(); cur = conn.cursor()
    cur.execute(f"UPDATE xodimlar SET {maydon}=%s WHERE id=%s", (qiymat, xodim_id))
    conn.commit(); cur.close(); conn.close()

def telegram_id_orqali_xodim(telegram_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id,ism,rol,kompaniya_id FROM xodimlar WHERE telegram_id=%s", (telegram_id,))
    r = cur.fetchone(); cur.close(); conn.close(); return r

def telefon_orqali_xodim(telefon):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id,ism,rol,kompaniya_id FROM xodimlar WHERE telefon LIKE %s", (f"%{telefon[-9:]}%",))
    r = cur.fetchone(); cur.close(); conn.close(); return r

def xodim_telegram_saqlash(xodim_id, telegram_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE xodimlar SET telegram_id=%s WHERE id=%s", (telegram_id, xodim_id))
    conn.commit(); cur.close(); conn.close()

def hr_idlari(komp_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT telegram_id FROM xodimlar WHERE kompaniya_id=%s AND rol='hr' AND telegram_id IS NOT NULL", (komp_id,))
    r = cur.fetchall(); cur.close(); conn.close(); return [x[0] for x in r]

def barcha_xodimlar_eslatma():
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT x.id,x.telegram_id,x.ism,x.ish_boshlanish,x.tugilgan_kun,k.id
                  FROM xodimlar x JOIN kompaniyalar k ON x.kompaniya_id=k.id
                  WHERE x.telegram_id IS NOT NULL AND x.holat='faol' AND k.holat='faol' ''')
    r = cur.fetchall(); cur.close(); conn.close(); return r

# ========== DAVOMAT ==========

def keldi_belgilash(xodim_id, komp_id, kiritdi="xodim", kiritdi_id=None):
    conn = connect(); cur = conn.cursor()
    sana = hozir().strftime("%Y-%m-%d")
    haqiqiy_vaqt = hozir().strftime("%H:%M")
    # Oxirgi yozuvni tekshir
    cur.execute("SELECT id, ketdi FROM davomat WHERE xodim_id=%s AND sana=%s ORDER BY id DESC LIMIT 1", (xodim_id, sana))
    row = cur.fetchone()
    if row:
        _, ketdi = row
        if not ketdi:
            cur.close(); conn.close(); return "already"
    cur.execute("SELECT ish_boshlanish FROM xodimlar WHERE id=%s", (xodim_id,))
    x = cur.fetchone(); kechikish = 0
    ish_bosh = x[0] if x else "08:00"
    # Ish vaqtidan oldin kelsa — ish boshlanish vaqtini yoz
    try:
        b = datetime.strptime(ish_bosh, "%H:%M")
        k = datetime.strptime(haqiqiy_vaqt, "%H:%M")
        vaqt = ish_bosh if k < b else haqiqiy_vaqt
        if k > b: kechikish = int((k - b).total_seconds() / 60)
    except:
        vaqt = haqiqiy_vaqt
    cur.execute('''INSERT INTO davomat(xodim_id,kompaniya_id,sana,keldi,kechikish,kiritdi,kiritdi_id)
                  VALUES(%s,%s,%s,%s,%s,%s,%s)''', (xodim_id,komp_id,sana,vaqt,kechikish,kiritdi,kiritdi_id))
    conn.commit(); cur.close(); conn.close()
    return f"keldi|{vaqt}|{kechikish}"

def keldi_rasm_saqlash(xodim_id, rasm_id):
    conn = connect(); cur = conn.cursor()
    sana = hozir().strftime("%Y-%m-%d")
    cur.execute("UPDATE davomat SET keldi_rasm=%s WHERE xodim_id=%s AND sana=%s", (rasm_id,xodim_id,sana))
    conn.commit(); cur.close(); conn.close()

def ketdi_belgilash(xodim_id, komp_id, kiritdi="xodim", kiritdi_id=None):
    conn = connect(); cur = conn.cursor()
    sana = hozir().strftime("%Y-%m-%d")
    haqiqiy_vaqt = hozir().strftime("%H:%M")
    cur.execute("SELECT id, keldi FROM davomat WHERE xodim_id=%s AND sana=%s AND ketdi IS NULL ORDER BY id DESC LIMIT 1", (xodim_id, sana))
    row = cur.fetchone()
    if not row or not row[1]: cur.close(); conn.close(); return "nokeldi"
    davomat_id, keldi_vaqt = row
    cur.execute("SELECT ish_tugash FROM xodimlar WHERE id=%s", (xodim_id,))
    x = cur.fetchone()
    ish_tugash = x[0] if x else "18:00"
    # Ish vaqtidan keyin ketsa — ish tugash vaqtini yoz
    try:
        tug = datetime.strptime(ish_tugash, "%H:%M")
        kv = datetime.strptime(haqiqiy_vaqt, "%H:%M")
        vaqt = ish_tugash if kv > tug else haqiqiy_vaqt
    except:
        vaqt = haqiqiy_vaqt
    try:
        k = datetime.strptime(keldi_vaqt, "%H:%M")
        kv2 = datetime.strptime(vaqt, "%H:%M")
        ish_soat = round((kv2 - k).total_seconds() / 3600, 2)
    except: ish_soat = 0
    cur.execute('''UPDATE davomat SET ketdi=%s,ish_soat=%s,kiritdi=%s,kiritdi_id=%s
                  WHERE id=%s''', (vaqt, ish_soat, kiritdi, kiritdi_id, davomat_id))
    conn.commit(); cur.close(); conn.close()
    return f"ketdi|{vaqt}|{ish_soat}|{ish_tugash}"

def ketdi_rasm_saqlash(xodim_id, rasm_id):
    conn = connect(); cur = conn.cursor()
    sana = hozir().strftime("%Y-%m-%d")
    cur.execute("UPDATE davomat SET ketdi_rasm=%s WHERE xodim_id=%s AND sana=%s", (rasm_id,xodim_id,sana))
    conn.commit(); cur.close(); conn.close()

def davomat_tahrirlash(davomat_id, maydon, qiymat):
    conn = connect(); cur = conn.cursor()
    cur.execute(f"UPDATE davomat SET {maydon}=%s WHERE id=%s", (qiymat, davomat_id))
    conn.commit(); cur.close(); conn.close()

def xodim_davomati(xodim_id, oy=None):
    """Get attendance records - TEXT DATA ONLY (no media/photos/videos)"""
    conn = connect(); cur = conn.cursor()
    if oy:
        # IMPORTANT: ONLY select text fields, NEVER select keldi_rasm or ketdi_rasm
        cur.execute('''SELECT id,sana,keldi,ketdi,ish_soat,kechikish,holat,izoh,kiritdi
                      FROM davomat WHERE xodim_id=%s AND sana LIKE %s ORDER BY sana''',
                    (xodim_id, f"%{oy}%"))
    else:
        # IMPORTANT: ONLY select text fields, NEVER select keldi_rasm or ketdi_rasm
        cur.execute('''SELECT id,sana,keldi,ketdi,ish_soat,kechikish,holat,izoh,kiritdi
                      FROM davomat WHERE xodim_id=%s ORDER BY sana DESC''', (xodim_id,))
    r = cur.fetchall(); cur.close(); conn.close(); return r

def xodim_davomati_text_only(xodim_id, oy=None):
    """Get attendance records - TEXT DATA ONLY (no media/photos/videos)
    CRITICAL SECURITY: Used by HR attendance viewing to ensure no rasm_id or video_id is shown"""
    conn = connect(); cur = conn.cursor()
    if oy:
        # CRITICAL SECURITY: ONLY select text fields, NEVER select keldi_rasm or ketdi_rasm
        cur.execute('''SELECT id,sana,keldi,ketdi,ish_soat,kechikish,holat
                      FROM davomat WHERE xodim_id=%s AND sana LIKE %s ORDER BY sana''',
                    (xodim_id, f"%{oy}%"))
    else:
        # CRITICAL SECURITY: ONLY select text fields, NEVER select keldi_rasm or ketdi_rasm
        cur.execute('''SELECT id,sana,keldi,ketdi,ish_soat,kechikish,holat
                      FROM davomat WHERE xodim_id=%s ORDER BY sana DESC''', (xodim_id,))
    r = cur.fetchall(); cur.close(); conn.close(); return r

def kompaniya_davomati(komp_id, sana=None):
    conn = connect(); cur = conn.cursor()
    if sana:
        cur.execute('''SELECT x.ism,d.sana,d.keldi,d.ketdi,d.ish_soat,d.kechikish,d.holat,d.id
                      FROM davomat d JOIN xodimlar x ON d.xodim_id=x.id
                      WHERE d.kompaniya_id=%s AND d.sana=%s ORDER BY x.ism''', (komp_id, sana))
    else:
        cur.execute('''SELECT x.ism,d.sana,d.keldi,d.ketdi,d.ish_soat,d.kechikish,d.holat,d.id
                      FROM davomat d JOIN xodimlar x ON d.xodim_id=x.id
                      WHERE d.kompaniya_id=%s ORDER BY d.sana DESC,x.ism''', (komp_id,))
    r = cur.fetchall(); cur.close(); conn.close(); return r

def manual_davomat(xodim_id, komp_id, sana, keldi_v, ketdi_v, holat, izoh, kiritdi_ism, kiritdi_id):
    conn = connect(); cur = conn.cursor()
    ish_soat = 0; kechikish = 0
    try:
        k = datetime.strptime(keldi_v, "%H:%M"); kt = datetime.strptime(ketdi_v, "%H:%M")
        ish_soat = round((kt - k).total_seconds() / 3600, 2)
        cur.execute("SELECT ish_boshlanish FROM xodimlar WHERE id=%s", (xodim_id,))
        x = cur.fetchone()
        if x:
            b = datetime.strptime(x[0], "%H:%M")
            if k > b: kechikish = int((k - b).total_seconds() / 60)
    except: pass
    cur.execute("SELECT id FROM davomat WHERE xodim_id=%s AND sana=%s", (xodim_id, sana))
    if cur.fetchone():
        cur.execute('''UPDATE davomat SET keldi=%s,ketdi=%s,ish_soat=%s,kechikish=%s,
                      holat=%s,izoh=%s,kiritdi=%s,kiritdi_id=%s WHERE xodim_id=%s AND sana=%s''',
                    (keldi_v,ketdi_v,ish_soat,kechikish,holat,izoh,kiritdi_ism,kiritdi_id,xodim_id,sana))
    else:
        cur.execute('''INSERT INTO davomat(xodim_id,kompaniya_id,sana,keldi,ketdi,ish_soat,
                      kechikish,holat,izoh,kiritdi,kiritdi_id)
                      VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
                    (xodim_id,komp_id,sana,keldi_v,ketdi_v,ish_soat,kechikish,holat,izoh,kiritdi_ism,kiritdi_id))
    conn.commit(); cur.close(); conn.close()
    return "✅ Davomat kiritildi!"

# ========== SABABLI SO'ROVLAR ==========

def sababli_sorov_saqlash(xodim_id, komp_id, sana, sabab):
    conn = connect(); cur = conn.cursor()
    cur.execute("INSERT INTO sababli_sorovlar(xodim_id,kompaniya_id,sana,sabab) VALUES(%s,%s,%s,%s) RETURNING id",
                (xodim_id,komp_id,sana,sabab))
    sid = cur.fetchone()[0]; conn.commit(); cur.close(); conn.close(); return sid

def sababli_sorov_yangilash(sorov_id, holat, xodim_id, sana, sabab):
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE sababli_sorovlar SET holat=%s WHERE id=%s", (holat, sorov_id))
    if holat == 'tasdiqlandi':
        cur.execute("SELECT id FROM davomat WHERE xodim_id=%s AND sana=%s", (xodim_id, sana))
        if cur.fetchone():
            cur.execute("UPDATE davomat SET holat='sababli',izoh=%s WHERE xodim_id=%s AND sana=%s",
                       (sabab,xodim_id,sana))
        else:
            cur.execute("INSERT INTO davomat(xodim_id,sana,holat,izoh) VALUES(%s,%s,'sababli',%s)",
                       (xodim_id,sana,sabab))
    conn.commit(); cur.close(); conn.close()

# ========== LIVE LOKATSIYA ==========

def live_lokatsiya_saqlash(xodim_id, komp_id, lat, lon):
    conn = connect(); cur = conn.cursor()
    vaqt = hozir().strftime("%H:%M:%S")
    cur.execute("SELECT id FROM live_lokatsiyalar WHERE xodim_id=%s AND faol=TRUE", (xodim_id,))
    if cur.fetchone():
        cur.execute("UPDATE live_lokatsiyalar SET lat=%s,lon=%s,vaqt=%s WHERE xodim_id=%s AND faol=TRUE",
                    (lat,lon,vaqt,xodim_id))
    else:
        cur.execute("INSERT INTO live_lokatsiyalar(xodim_id,kompaniya_id,lat,lon,vaqt,faol) VALUES(%s,%s,%s,%s,%s,TRUE)",
                    (xodim_id,komp_id,lat,lon,vaqt))
    conn.commit(); cur.close(); conn.close()

def live_lokatsiya_olish(xodim_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT lat,lon,vaqt FROM live_lokatsiyalar WHERE xodim_id=%s AND faol=TRUE", (xodim_id,))
    r = cur.fetchone(); cur.close(); conn.close()
    if not r: return None
    # So'nggi yangilanish 10 daqiqadan eski bo'lsa — o'chgan
    try:
        from datetime import datetime
        last = datetime.strptime(r[2], "%H:%M:%S")
        now = hozir()
        last_full = now.replace(hour=last.hour, minute=last.minute, second=last.second)
        diff = (now - last_full).total_seconds()
        if diff > 600:  # 10 daqiqa
            return None
    except: pass
    return r

def live_lokatsiya_ochirish(xodim_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE live_lokatsiyalar SET faol=FALSE WHERE xodim_id=%s", (xodim_id,))
    conn.commit(); cur.close(); conn.close()

def barcha_live_xodimlar():
    conn = connect(); cur = conn.cursor()
    bugun = hozir().strftime("%Y-%m-%d")
    cur.execute('''SELECT ll.xodim_id,ll.kompaniya_id,ll.lat,ll.lon,x.ism,x.telegram_id,
                  k.gps_lat,k.gps_lon,k.gps_radius,k.admin_id,k.nomi
                  FROM live_lokatsiyalar ll
                  JOIN xodimlar x ON ll.xodim_id=x.id
                  JOIN kompaniyalar k ON ll.kompaniya_id=k.id
                  WHERE ll.faol=TRUE AND k.live_gps_tekshiruv=TRUE AND k.holat='faol'
                  AND EXISTS(SELECT 1 FROM davomat d WHERE d.xodim_id=ll.xodim_id
                             AND d.sana=%s AND d.keldi IS NOT NULL AND d.ketdi IS NULL)''', (bugun,))
    r = cur.fetchall(); cur.close(); conn.close(); return r

# ========== STATISTIKA ==========

def xodim_oy_statistika(xodim_id, oy=None):
    conn = connect(); cur = conn.cursor()
    if not oy: oy = hozir().strftime("%Y-%m")
    cur.execute('''SELECT COUNT(*),COALESCE(SUM(ish_soat),0),
                  COUNT(CASE WHEN kechikish>0 THEN 1 END),
                  COALESCE(SUM(kechikish),0),
                  COUNT(CASE WHEN holat='sababli' THEN 1 END),
                  COUNT(CASE WHEN keldi IS NULL THEN 1 END),
                  COUNT(CASE WHEN holat='sababsiz' THEN 1 END)
                  FROM davomat WHERE xodim_id=%s AND sana LIKE %s''',
                (xodim_id, f"{oy}%"))
    r = cur.fetchone(); cur.close(); conn.close(); return r

def xodim_oylik_hisob(xodim_id, oy=None):
    """Oylik ish haqi — KPI sozlamalari bilan hisoblash"""
    if not oy: oy = hozir().strftime("%Y-%m")
    conn = connect(); cur = conn.cursor()

    # Asosiy oylik va davomat statistikasi bitta so'rovda
    cur.execute('''SELECT x.oylik,
                   COUNT(CASE WHEN d.holat='sababsiz' THEN 1 END),
                   COUNT(CASE WHEN d.holat='sababli'  THEN 1 END),
                   COUNT(CASE WHEN d.kechikish > 0    THEN 1 END),
                   COALESCE(SUM(d.kechikish), 0)
                   FROM xodimlar x
                   LEFT JOIN davomat d ON d.xodim_id=x.id AND d.sana LIKE %s
                   WHERE x.id=%s GROUP BY x.oylik''', (f"{oy}%", xodim_id))
    row = cur.fetchone()

    # KPI sozlamalari
    cur.execute("SELECT * FROM xodim_kpi WHERE xodim_id=%s", (xodim_id,))
    kpi = cur.fetchone()
    cur.close(); conn.close()

    if not row or not row[0]: return 0, 0, 0
    oylik = float(row[0])
    sababsiz      = row[1] or 0
    sababli       = row[2] or 0
    kech_kun      = row[3] or 0

    # Kunlik ish haqi
    yil, oy_raqam = int(oy[:4]), int(oy[5:])
    _, kun_soni = calendar.monthrange(yil, oy_raqam)
    ish_kunlari = sum(1 for d in range(1, kun_soni+1) if date(yil, oy_raqam, d).weekday() < 5)
    kunlik = oylik / ish_kunlari if ish_kunlari > 0 else 0

    # KPI jarimalarini qo'llash
    if kpi:
        sababsiz_jarima = float(kpi[5] or 0) * sababsiz   # sababsiz_qolish_jarima
        sababli_jarima  = float(kpi[7] or 0) * sababli    # sababli_qolish_jarima
        kech_jarima     = float(kpi[9] or 0) * kech_kun   # kech_kelish_jarima
        qosh_jarima     = float(kpi[15] or 0)             # qoshimcha_jarima
    else:
        # KPI yo'q bo'lsa — kunlik stavkadan hisoblash
        sababsiz_jarima = kunlik * sababsiz
        sababli_jarima  = kunlik * sababli
        kech_jarima     = 0
        qosh_jarima     = 0

    jami_jarima = sababsiz_jarima + sababli_jarima + kech_jarima + qosh_jarima
    hisoblangan = max(0, oylik - jami_jarima)
    return hisoblangan, jami_jarima, sababsiz_jarima

def kompaniya_reyting(komp_id, oy=None):
    conn = connect(); cur = conn.cursor()
    if not oy: oy = hozir().strftime("%Y-%m")
    cur.execute('''SELECT x.ism,x.lavozim,COUNT(d.id),
                  COALESCE(SUM(d.ish_soat),0),COALESCE(SUM(d.kechikish),0),
                  COUNT(CASE WHEN d.kechikish>0 THEN 1 END)
                  FROM xodimlar x LEFT JOIN davomat d ON x.id=d.xodim_id AND d.sana LIKE %s
                  WHERE x.kompaniya_id=%s AND x.holat='faol'
                  GROUP BY x.id,x.ism,x.lavozim ORDER BY 5 ASC,3 DESC''',
                (f"{oy}%", komp_id))
    r = cur.fetchall(); cur.close(); conn.close(); return r

def haftalik_davomat_kompaniyalar():
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id,nomi,admin_id FROM kompaniyalar WHERE holat='faol' AND admin_id IS NOT NULL")
    r = cur.fetchall(); cur.close(); conn.close(); return r

def kompaniya_haftalik_stat(komp_id):
    conn = connect(); cur = conn.cursor()
    bugun = hozir().date()
    dushanba = bugun - timedelta(days=bugun.weekday())
    juma = dushanba + timedelta(days=4)
    cur.execute('''SELECT COUNT(DISTINCT xodim_id),COUNT(DISTINCT sana),
                  COALESCE(SUM(ish_soat),0),COALESCE(AVG(CASE WHEN kechikish>0 THEN kechikish END),0)
                  FROM davomat WHERE kompaniya_id=%s AND sana>=%s AND sana<=%s''',
                (komp_id, str(dushanba), str(juma)))
    r = cur.fetchone(); cur.close(); conn.close(); return r

# ========== HISOBOTLAR ==========

def super_admin_hisobot():
    conn = connect(); cur = conn.cursor()
    wb = openpyxl.Workbook()

    ws1 = wb.active; ws1.title = "Kompaniyalar"
    ws1.append(["#","Kompaniya","Admin telefon","Holat","Xodimlar","Yaratilgan"])
    cur.execute("SELECT id,nomi,admin_telefon,holat,yaratilgan FROM kompaniyalar ORDER BY id")
    kompaniyalar = cur.fetchall()
    for i,k in enumerate(kompaniyalar,1):
        cur.execute("SELECT COUNT(*) FROM xodimlar WHERE kompaniya_id=%s", (k[0],))
        ws1.append([i,k[1],k[2],k[3],cur.fetchone()[0],k[4]])

    ws2 = wb.create_sheet("Xodimlar")
    ws2.append(["#","Kompaniya","Ism","Lavozim","Telefon","Oylik","Ish vaqti","Rol","Holat"])
    cur.execute('''SELECT k.nomi,x.ism,x.lavozim,x.telefon,x.oylik,x.ish_boshlanish,x.ish_tugash,x.rol,x.holat
                  FROM xodimlar x JOIN kompaniyalar k ON x.kompaniya_id=k.id ORDER BY k.id,x.id''')
    for i,x in enumerate(cur.fetchall(),1):
        ws2.append([i,x[0],x[1],x[2],x[3],x[4],f"{x[5]}-{x[6]}",x[7],x[8]])

    ws3 = wb.create_sheet("Davomat")
    ws3.append(["#","Kompaniya","Xodim","Sana","Keldi","Ketdi","Ish vaqti","Kechikish","Holat"])
    cur.execute('''SELECT k.nomi,x.ism,d.sana,d.keldi,d.ketdi,d.ish_soat,d.kechikish,d.holat
                  FROM davomat d JOIN xodimlar x ON d.xodim_id=x.id
                  JOIN kompaniyalar k ON d.kompaniya_id=k.id ORDER BY k.id,d.sana''')
    for i,d in enumerate(cur.fetchall(),1):
        ws3.append([i,d[0],d[1],d[2],d[3],d[4],soat_format(d[5]),kechikish_format(d[6]),d[7]])

    ws4 = wb.create_sheet("Statistika")
    ws4.append(["Kompaniya","Jami xodim","Bugun kelgan","O'rt kechikish","Holat"])
    bugun = hozir().strftime("%Y-%m-%d")
    for k in kompaniyalar:
        cur.execute("SELECT COUNT(*) FROM xodimlar WHERE kompaniya_id=%s", (k[0],))
        jami = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM davomat WHERE kompaniya_id=%s AND sana=%s", (k[0],bugun))
        bk = cur.fetchone()[0]
        cur.execute("SELECT AVG(kechikish) FROM davomat WHERE kompaniya_id=%s", (k[0],))
        ort = cur.fetchone()[0] or 0
        ws4.append([k[1],jami,bk,round(ort,1),k[3]])

    cur.close(); conn.close()
    wb.save("super_hisobot.xlsx")
    return "super_hisobot.xlsx"

def kompaniya_hisobot(komp_id):
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT nomi FROM kompaniyalar WHERE id=%s", (komp_id,))
    r = cur.fetchone()
    if not r: return None
    nomi = r[0]; wb = openpyxl.Workbook()

    ws1 = wb.active; ws1.title = "Xodimlar"
    ws1.append(["#","Ism","Lavozim","Telefon","Oylik","Ish vaqti","Rol","Holat"])
    cur.execute('''SELECT id,ism,lavozim,telefon,oylik,ish_boshlanish,ish_tugash,rol,holat
                  FROM xodimlar WHERE kompaniya_id=%s ORDER BY id''', (komp_id,))
    xodimlar = cur.fetchall()
    for i,x in enumerate(xodimlar,1):
        ws1.append([i,x[1],x[2],x[3],x[4],f"{x[5]}-{x[6]}",x[7],x[8]])

    ws2 = wb.create_sheet("Davomat")
    ws2.append(["#","Xodim","Sana","Keldi","Ketdi","Ish vaqti","Kechikish","Holat","Izoh"])
    cur.execute('''SELECT x.ism,d.sana,d.keldi,d.ketdi,d.ish_soat,d.kechikish,d.holat,d.izoh
                  FROM davomat d JOIN xodimlar x ON d.xodim_id=x.id
                  WHERE d.kompaniya_id=%s ORDER BY d.sana DESC,x.ism''', (komp_id,))
    for i,d in enumerate(cur.fetchall(),1):
        ws2.append([i,d[0],d[1],d[2],d[3],soat_format(d[4]),kechikish_format(d[5]),d[6],d[7] or "—"])

    ws3 = wb.create_sheet("Statistika")
    ws3.append(["Xodim","Lavozim","Jami kun","Jami ish soat","O'rt kechikish","Kech kelgan"])
    for x in xodimlar:
        cur.execute('''SELECT COUNT(*),SUM(ish_soat),AVG(kechikish),COUNT(CASE WHEN kechikish>0 THEN 1 END)
                      FROM davomat WHERE xodim_id=%s''', (x[0],))
        s = cur.fetchone()
        ws3.append([x[1],x[2],s[0] or 0,soat_format(s[1] or 0),kechikish_format(s[2] or 0),s[3] or 0])

    cur.close(); conn.close()
    fayl = f"hisobot_{nomi}.xlsx"
    wb.save(fayl); return fayl

def kompaniya_hisobot_kunlik(komp_id):
    """Kunlik batafsil hisobot - har bir kun uchun alohida ustun"""
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT nomi FROM kompaniyalar WHERE id=%s", (komp_id,))
    r = cur.fetchone()
    if not r: return None
    nomi = r[0]

    # Bugungi oyni aniqlash
    bugun = hozir()
    oy_boshi = bugun.replace(day=1)

    # Xodimlarni olish
    cur.execute('''SELECT id,ism,lavozim FROM xodimlar
                  WHERE kompaniya_id=%s ORDER BY ism''', (komp_id,))
    xodimlar = cur.fetchall()

    # Excel workbook yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Davomat"

    # Header - kompaniya nomi va sana
    ws['A1'] = f"Taskiloti: {nomi}"
    ws['A2'] = f"Sana: {bugun.strftime('%d.%m.%Y')}"

    # Ustun sarlavhalari
    headers = ["Taskiloti", "Sana", "Xodim ismi", "Lavozimi"]

    # Oy ichidagi barcha kunlar uchun ustunlar qo'shish
    kun_raqami = 1
    kun_dict = {}  # {kun_raqami: column_index}

    current_date = oy_boshi
    col_idx = 5  # A=1, B=2, C=3, D=4, E=5

    while current_date.month == bugun.month and current_date <= bugun:
        kun_sarlavha = current_date.strftime('%d.%m')
        headers.append(f"Keldi\n{kun_sarlavha}")
        headers.append(f"Ketdi\n{kun_sarlavha}")
        kun_dict[current_date.date()] = col_idx
        col_idx += 2
        current_date += timedelta(days=1)

    # Headers yozish
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=header)

    # Xodimlar va ularning davomati
    row = 5
    for xodim_id, ism, lavozim in xodimlar:
        ws.cell(row=row, column=1, value=nomi)
        ws.cell(row=row, column=2, value=bugun.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=3, value=ism)
        ws.cell(row=row, column=4, value=lavozim)

        # Davomat ma'lumotlarini olish
        cur.execute('''SELECT sana,keldi,ketdi,holat FROM davomat
                      WHERE xodim_id=%s AND sana >= %s AND sana <= %s
                      ORDER BY sana''', (xodim_id, oy_boshi.strftime('%Y-%m-%d'), bugun.strftime('%Y-%m-%d')))
        davomatlar = {d[0]: d for d in cur.fetchall()}

        # Har bir kun uchun keldi/ketdi vaqtini yozish
        current_date = oy_boshi
        while current_date.month == bugun.month and current_date <= bugun:
            sana = current_date.date()
            col_idx = kun_dict[sana]

            if sana in davomatlar:
                _, keldi, ketdi, holat = davomatlar[sana]
                ws.cell(row=row, column=col_idx, value=keldi or "—")
                ws.cell(row=row, column=col_idx+1, value=ketdi or "—")
            else:
                # Agar davomat yo'q bo'lsa, bo'sh yoki holat ko'rsatish
                ws.cell(row=row, column=col_idx, value="—")
                ws.cell(row=row, column=col_idx+1, value="—")

            current_date += timedelta(days=1)

        row += 1

    cur.close(); conn.close()
    fayl = f"hisobot_kunlik_{nomi}_{bugun.strftime('%Y%m%d')}.xlsx"
    wb.save(fayl)
    return fayl

def hisobot_row_format(komp_id=None, sana_from=None, sana_to=None, super_admin=False):
    """Row-based detailed hisobot - har bir kun alohida qator
    sana_from, sana_to: YYYY-MM-DD format strings
    komp_id: None if super_admin=True (all companies), else specific company
    """
    conn = connect(); cur = conn.cursor()
    bugun = hozir()

    # Excel workbook yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Davomat"

    # Title va sana
    title = "BARCHA KOMPANIYALAR DAVOMATI" if super_admin else "DAVOMAT HISOBOTI"
    ws['A1'] = title
    ws['A2'] = f"SANA ORALIG'I: {sana_from} dan {sana_to} gacha"

    # CAPITAL LETTER headers - yangi YARATGAN ustuni qo'shildi
    headers = [
        "TASKILOTI", "SANA", "XODIM ISM", "LAVOZIMI",
        "KELDI", "KETDI", "ISH SOAT", "KECHIKISH", "HOLAT", "IZOH", "YARATGAN"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Davomat ma'lumotlarini olish - kiritdi va kiritdi_id qo'shildi
    if super_admin:
        cur.execute('''SELECT k.nomi, d.sana, x.ism, x.lavozim,
                              d.keldi, d.ketdi, d.ish_soat, d.kechikish, d.holat, d.izoh,
                              d.kiritdi, d.kiritdi_id
                       FROM davomat d
                       JOIN xodimlar x ON d.xodim_id=x.id
                       JOIN kompaniyalar k ON d.kompaniya_id=k.id
                       WHERE d.sana >= %s AND d.sana <= %s
                       ORDER BY k.nomi, d.sana, x.ism''',
                    (sana_from, sana_to))
    else:
        cur.execute('''SELECT k.nomi, d.sana, x.ism, x.lavozim,
                              d.keldi, d.ketdi, d.ish_soat, d.kechikish, d.holat, d.izoh,
                              d.kiritdi, d.kiritdi_id
                       FROM davomat d
                       JOIN xodimlar x ON d.xodim_id=x.id
                       JOIN kompaniyalar k ON d.kompaniya_id=k.id
                       WHERE d.kompaniya_id=%s AND d.sana >= %s AND d.sana <= %s
                       ORDER BY d.sana, x.ism''',
                    (komp_id, sana_from, sana_to))

    davomatlar = cur.fetchall()

    from datetime import date, timedelta

    # OPTIMIZED: Barcha ma'lumotni BIR MARTA olish
    if super_admin:
        cur.execute('''SELECT DISTINCT xodim_id, x.ism, x.lavozim, k.nomi
                       FROM davomat d
                       JOIN xodimlar x ON d.xodim_id=x.id
                       JOIN kompaniyalar k ON d.kompaniya_id=k.id
                       WHERE d.sana >= %s AND d.sana <= %s
                       ORDER BY k.nomi, x.ism''', (sana_from, sana_to))
    else:
        cur.execute('''SELECT DISTINCT d.xodim_id, x.ism, x.lavozim, k.nomi
                       FROM davomat d
                       JOIN xodimlar x ON d.xodim_id=x.id
                       JOIN kompaniyalar k ON d.kompaniya_id=k.id
                       WHERE d.kompaniya_id=%s AND d.sana >= %s AND d.sana <= %s
                       ORDER BY x.ism''', (komp_id, sana_from, sana_to))

    xodimlar_list = cur.fetchall()

    # Barcha davomatlarni BIR MARTA olish (date dict sifatida)
    if super_admin:
        cur.execute('''SELECT d.xodim_id, d.sana, d.keldi, d.ketdi, d.ish_soat, d.kechikish,
                              d.holat, d.izoh, d.kiritdi, d.kiritdi_id
                       FROM davomat d
                       WHERE d.sana >= %s AND d.sana <= %s
                       ORDER BY d.xodim_id, d.sana''', (sana_from, sana_to))
    else:
        cur.execute('''SELECT d.xodim_id, d.sana, d.keldi, d.ketdi, d.ish_soat, d.kechikish,
                              d.holat, d.izoh, d.kiritdi, d.kiritdi_id
                       FROM davomat d
                       WHERE d.kompaniya_id=%s AND d.sana >= %s AND d.sana <= %s
                       ORDER BY d.xodim_id, d.sana''', (komp_id, sana_from, sana_to))

    # Dictionary sifatida saqlash: {xodim_id: {sana: (keldi, ketdi, ...)}
    davomatlar_dict = {}
    for xodim_id, sana, keldi, ketdi, ish_soat, kechikish, holat, izoh, kiritdi, kiritdi_id in cur.fetchall():
        if xodim_id not in davomatlar_dict:
            davomatlar_dict[xodim_id] = {}
        davomatlar_dict[xodim_id][sana] = (keldi, ketdi, ish_soat, kechikish, holat, izoh, kiritdi, kiritdi_id)

    # Admin ismlarini cache qilish
    admin_names_cache = {}

    def get_admin_name(kiritdi_id, kiritdi_type, xodim_id_for_default):
        """Admin/HR/Super Admin ismini cache bilan olish"""
        if kiritdi_id in admin_names_cache:
            return admin_names_cache[kiritdi_id]

        conn2 = connect()
        cur2 = conn2.cursor()
        cur2.execute("SELECT ism FROM super_adminlar WHERE telegram_id=%s", (kiritdi_id,))
        result = cur2.fetchone()
        if result:
            admin_names_cache[kiritdi_id] = result[0]
            cur2.close(); conn2.close()
            return result[0]

        cur2.execute("SELECT ism FROM xodimlar WHERE telegram_id=%s", (kiritdi_id,))
        result = cur2.fetchone()
        admin_names_cache[kiritdi_id] = result[0] if result else kiritdi_type.upper()
        cur2.close(); conn2.close()
        return admin_names_cache[kiritdi_id]

    # Data yozish - Python da iteratsiya (database sorash yo'q!)
    row = 5
    start_date = datetime.strptime(sana_from, "%Y-%m-%d").date()
    end_date = datetime.strptime(sana_to, "%Y-%m-%d").date()

    for xodim_id, xodim_ism, lavozim, komp_nomi in xodimlar_list:
        current_date = start_date

        while current_date <= end_date:
            sana_str = current_date.strftime("%Y-%m-%d")

            # Dictionary dan qidirish (database sorash yo'q!)
            dav = None
            if xodim_id in davomatlar_dict and sana_str in davomatlar_dict[xodim_id]:
                dav = davomatlar_dict[xodim_id][sana_str]

            # Kirituvchi ismini olish
            yaratgan = ""
            if dav:
                keldi, ketdi, ish_soat, kechikish, holat, izoh, kiritdi, kiritdi_id = dav

                if kiritdi == 'bot':
                    yaratgan = "BOT"
                elif kiritdi == 'xodim':
                    yaratgan = f"xodim + {xodim_ism}"
                elif kiritdi in ['admin', 'super_admin', 'hr']:
                    user_ism = get_admin_name(kiritdi_id, kiritdi, xodim_id)
                    yaratgan = f"{kiritdi} + {user_ism}"
            else:
                # Ma'lumot yo'q
                keldi = ketdi = ish_soat = kechikish = holat = izoh = None

            # Excel da yozish
            ws.cell(row=row, column=1, value=komp_nomi)
            ws.cell(row=row, column=2, value=sana_str)
            ws.cell(row=row, column=3, value=xodim_ism)
            ws.cell(row=row, column=4, value=lavozim)
            ws.cell(row=row, column=5, value=keldi or "—")
            ws.cell(row=row, column=6, value=ketdi or "—")
            ws.cell(row=row, column=7, value=soat_format(ish_soat) if ish_soat else "—")
            ws.cell(row=row, column=8, value=kechikish_format(kechikish) if kechikish else "—")
            ws.cell(row=row, column=9, value=holat or "—")
            ws.cell(row=row, column=10, value=izoh or "—")
            ws.cell(row=row, column=11, value=yaratgan)

            row += 1
            current_date += timedelta(days=1)

    cur.close(); conn.close()

    # Filename
    if super_admin:
        fayl = f"hisobot_{sana_from}_{sana_to}.xlsx"
    else:
        cur2 = connect().cursor()
        cur2.execute("SELECT nomi FROM kompaniyalar WHERE id=%s", (komp_id,))
        komp_data = cur2.fetchone()
        komp_name = komp_data[0] if komp_data else "hisobot"
        cur2.close()
        fayl = f"hisobot_{komp_name}_{sana_from}_{sana_to}.xlsx"

    wb.save(fayl)
    return fayl

def super_admin_hisobot_kunlik():
    """Super admin uchun barcha kompaniyalarning kunlik hisoboti"""
    conn = connect(); cur = conn.cursor()
    bugun = hozir()
    oy_boshi = bugun.replace(day=1)

    # Barcha kompaniyalarni olish
    cur.execute("SELECT id,nomi FROM kompaniyalar WHERE holat='faol' ORDER BY nomi")
    kompaniyalar = cur.fetchall()

    # Excel workbook yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Davomat"

    ws['A1'] = "BARCHA KOMPANIYALAR DAVOMATI"
    ws['A2'] = f"Sana: {bugun.strftime('%d.%m.%Y')}"

    # Ustun sarlavhalari
    headers = ["Taskiloti", "Sana", "Xodim ismi", "Lavozimi"]

    # Oy ichidagi barcha kunlar uchun ustunlar
    kun_dict = {}
    current_date = oy_boshi
    col_idx = 5

    while current_date.month == bugun.month and current_date <= bugun:
        kun_sarlavha = current_date.strftime('%d.%m')
        headers.append(f"Keldi\n{kun_sarlavha}")
        headers.append(f"Ketdi\n{kun_sarlavha}")
        kun_dict[current_date.date()] = col_idx
        col_idx += 2
        current_date += timedelta(days=1)

    # Headers yozish
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=header)

    # Barcha xodimlar bo'yicha
    row = 5
    for komp_id, komp_nomi in kompaniyalar:
        cur.execute('''SELECT id,ism,lavozim FROM xodimlar
                      WHERE kompaniya_id=%s ORDER BY ism''', (komp_id,))
        xodimlar = cur.fetchall()

        for xodim_id, ism, lavozim in xodimlar:
            ws.cell(row=row, column=1, value=komp_nomi)
            ws.cell(row=row, column=2, value=bugun.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=3, value=ism)
            ws.cell(row=row, column=4, value=lavozim)

            # Davomat ma'lumotlarini olish
            cur.execute('''SELECT sana,keldi,ketdi FROM davomat
                          WHERE xodim_id=%s AND sana >= %s AND sana <= %s
                          ORDER BY sana''', (xodim_id, oy_boshi.strftime('%Y-%m-%d'), bugun.strftime('%Y-%m-%d')))
            davomatlar = {d[0]: d for d in cur.fetchall()}

            # Har bir kun uchun keldi/ketdi
            current_date = oy_boshi
            while current_date.month == bugun.month and current_date <= bugun:
                sana = current_date.date()
                col_idx = kun_dict[sana]

                if sana in davomatlar:
                    _, keldi, ketdi = davomatlar[sana]
                    ws.cell(row=row, column=col_idx, value=keldi or "—")
                    ws.cell(row=row, column=col_idx+1, value=ketdi or "—")
                else:
                    ws.cell(row=row, column=col_idx, value="—")
                    ws.cell(row=row, column=col_idx+1, value="—")

                current_date += timedelta(days=1)

            row += 1

    cur.close(); conn.close()
    fayl = f"hisobot_barcha_{bugun.strftime('%Y%m%d')}.xlsx"
    wb.save(fayl)
    return fayl

# ========== AUDIT LOG ==========

def audit_log_qoshish(komp_id, amal, tafsilot, xodim_id=None, rasm_id=None, video_id=None, user_id=None, user_ism=None):
    """Audit log'ga amal qo'shish"""
    conn = connect(); cur = conn.cursor()
    vaqt = hozir().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute('''INSERT INTO audit_log(kompaniya_id,xodim_id,amal,tafsilot,vaqt,rasm_id,video_id,user_id,user_ism)
                  VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
                (komp_id, xodim_id, amal, tafsilot, vaqt, rasm_id, video_id, user_id, user_ism))
    conn.commit(); cur.close(); conn.close()

def audit_log_olish(komp_id, limit=50, offset=0):
    """Kompaniya'ning audit log'larini olish"""
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT id,xodim_id,amal,tafsilot,vaqt,rasm_id,video_id,user_ism
                  FROM audit_log WHERE kompaniya_id=%s
                  ORDER BY vaqt DESC LIMIT %s OFFSET %s''',
                (komp_id, limit, offset))
    r = cur.fetchall(); cur.close(); conn.close(); return r

def super_admin_audit_log(limit=100, offset=0):
    """Super admin - barcha kompaniya audit log'lari"""
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT k.nomi,al.amal,al.tafsilot,al.vaqt,al.user_ism,al.xodim_id,x.ism
                  FROM audit_log al
                  JOIN kompaniyalar k ON al.kompaniya_id=k.id
                  LEFT JOIN xodimlar x ON al.xodim_id=x.id
                  ORDER BY al.vaqt DESC LIMIT %s OFFSET %s''',
                (limit, offset))
    r = cur.fetchall(); cur.close(); conn.close(); return r

def export_audit_log_excel(komp_id=None):
    """Audit log'ni Excel'ga export qilish"""
    conn = connect(); cur = conn.cursor()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Log"
    ws.append(["#", "Kompaniya", "Amal", "Xodim", "Vaqt", "User", "Tafsilot"])

    if komp_id:
        cur.execute('''SELECT al.amal,k.nomi,x.ism,al.vaqt,al.user_ism,al.tafsilot
                      FROM audit_log al
                      JOIN kompaniyalar k ON al.kompaniya_id=k.id
                      LEFT JOIN xodimlar x ON al.xodim_id=x.id
                      WHERE al.kompaniya_id=%s ORDER BY al.vaqt DESC''', (komp_id,))
    else:
        cur.execute('''SELECT al.amal,k.nomi,x.ism,al.vaqt,al.user_ism,al.tafsilot
                      FROM audit_log al
                      JOIN kompaniyalar k ON al.kompaniya_id=k.id
                      LEFT JOIN xodimlar x ON al.xodim_id=x.id
                      ORDER BY al.vaqt DESC''')

    for i, row in enumerate(cur.fetchall(), 1):
        amal, komp_nomi, xodim_ism, vaqt, user_ism, tafsilot = row
        ws.append([i, komp_nomi, amal, xodim_ism or '—', vaqt, user_ism, tafsilot[:50]])

    cur.close(); conn.close()
    fayl = f"audit_log_{komp_id or 'all'}.xlsx"
    wb.save(fayl); return fayl

# ========== MOTIVATSIYA TIZIMI ==========

def xodim_streak_olish(xodim_id):
    """Xodimning davomiylik streakini olish"""
    conn = connect(); cur = conn.cursor()
    bugun = hozir().date()
    streak = 0

    for i in range(30):
        check_date = (bugun - timedelta(days=i)).strftime("%Y-%m-%d")
        cur.execute("SELECT id FROM davomat WHERE xodim_id=%s AND sana=%s AND keldi IS NOT NULL", (xodim_id, check_date))
        if cur.fetchone(): streak += 1
        else: break

    cur.close(); conn.close(); return streak

def haftalik_reyting_xodimlar(komp_id):
    """Kompaniyaning haftalik top 5 xodimi"""
    conn = connect(); cur = conn.cursor()
    bugun = hozir().date()
    dushanba = bugun - timedelta(days=bugun.weekday())
    juma = dushanba + timedelta(days=4)

    cur.execute('''SELECT x.id,x.ism,x.lavozim,
                  COUNT(DISTINCT d.sana) as kun,
                  COALESCE(SUM(d.ish_soat),0) as soat,
                  COUNT(CASE WHEN d.kechikish>0 THEN 1 END) as kechikkan,
                  COALESCE(AVG(d.kechikish),0) as ort_kechikish
                  FROM xodimlar x
                  LEFT JOIN davomat d ON x.id=d.xodim_id
                  WHERE x.kompaniya_id=%s AND x.holat='faol'
                  AND (d.sana IS NULL OR (d.sana>=%s AND d.sana<=%s AND d.holat NOT IN ('kasal','ta\'til')))
                  GROUP BY x.id,x.ism,x.lavozim
                  ORDER BY ort_kechikish ASC, kun DESC, soat DESC
                  LIMIT 5''', (komp_id, str(dushanba), str(juma)))
    r = cur.fetchall(); cur.close(); conn.close(); return r

def xodim_bugun_statistika(xodim_id):
    """Bugungi xodim statistikasi"""
    conn = connect(); cur = conn.cursor()
    bugun = hozir().strftime("%Y-%m-%d")
    cur.execute('''SELECT keldi,ketdi,ish_soat,kechikish,holat
                  FROM davomat WHERE xodim_id=%s AND sana=%s''', (xodim_id, bugun))
    r = cur.fetchone(); cur.close(); conn.close(); return r

def generate_keldi_motivation(xodim, kechikish, streak):
    """Keldi uchun motivatsiya matni"""
    ism = xodim[1] if xodim else 'Xodim'
    if kechikish == 0:
        msg = f"🎉 *{ism}! Vaqtida keldingiz!*\n"
        if streak >= 5: msg += f"🔥 {streak} kunlik streak! Zo'r!"
        elif streak >= 3: msg += f"💪 {streak} kunlik davomiylik! Yaxshi!"
        return msg
    else:
        kech_fmt = kechikish_format(kechikish)
        return f"⚠️ *{ism}!* Bugun {kech_fmt}ga kechiktingiz.\nKeying safar vaqtida keling!"

def generate_ketdi_motivation(xodim, ish_tugash, ketdi_vaqt, ish_soat, streak):
    """Ketdi uchun motivatsiya matni"""
    ism = xodim[1] if xodim else 'Xodim'
    try:
        tug = datetime.strptime(ish_tugash, "%H:%M")
        ket = datetime.strptime(ketdi_vaqt, "%H:%M")
        farq = int((ket - tug).total_seconds() / 60)
    except:
        farq = 0

    s = int(ish_soat); d = int((ish_soat - s) * 60)
    farq_fmt = kechikish_format(abs(farq))

    if farq >= 30:
        return (f"🌙 *QOLIB ISHLASH!* (+{farq_fmt} ortiqcha)\n"
                f"🏆 Rahmat, {ism}! Bugun {s} soat {d} daqiqa ishlading.\n"
                f"Qo'shimcha mehnat ko'rishingiz qadrlanadi! 💪\n"
                f"{'🔥 ' + str(streak) + ' kunlik streak!' if streak >= 3 else ''}")
    elif farq >= -15:
        return (f"✅ *TO'LIQ ISH KUNI!*\n"
                f"Yaxshi ish, {ism}!\n⏱ {s} soat {d} daqiqa samarali mehnat. 💼\n"
                f"{'🔥 Streakingiz: ' + str(streak) + ' kun!' if streak >= 3 else ''}")
    elif farq >= -30:
        return (f"⚠️ *ERTA KETISH* ({farq_fmt} oldin)\n"
                f"{ism}, ish vaqti tugamay ketdingiz.\n"
                f"Sababini HR ga bildiring! 📋")
    else:
        return (f"❌ *JIDDIY ERTA KETISH!* ({farq_fmt} oldin)\n"
                f"Bu intizom buzilishi! HR va admin xabardor qilindi! 🚫")

def xodim_statistika(xodim_id, sana_1, sana_2):
    """Calculate attendance statistics for a date range"""
    conn = connect(); cur = conn.cursor()

    # Get all days in range
    sana1 = datetime.strptime(sana_1, "%Y-%m-%d")
    sana2 = datetime.strptime(sana_2, "%Y-%m-%d")
    days_in_range = (sana2 - sana1).days + 1

    # Get attendance data
    cur.execute('''SELECT COUNT(*),
                   SUM(CASE WHEN holat='normal' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN holat='sabsiz' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN holat='sababli' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN holat='kasal' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN holat='ta\'til' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN kechikish > 0 THEN 1 ELSE 0 END),
                   AVG(CASE WHEN kechikish > 0 THEN kechikish ELSE NULL END)
                FROM davomat
                WHERE xodim_id=%s AND sana BETWEEN %s AND %s''',
                (xodim_id, sana_1, sana_2))
    r = cur.fetchone()

    total_entries = r[0] or 0
    normal_days = r[1] or 0
    absent_days = r[2] or 0
    reason_days = r[3] or 0
    sick_days = r[4] or 0
    vacation_days = r[5] or 0
    late_days = r[6] or 0
    avg_delay = r[7] or 0

    cur.close(); conn.close()

    return {
        'days_in_range': days_in_range,
        'total_entries': total_entries,
        'normal_days': normal_days,
        'absent_days': absent_days,
        'reason_days': reason_days,
        'sick_days': sick_days,
        'vacation_days': vacation_days,
        'late_days': late_days,
        'avg_delay': avg_delay
    }

# ========== XABARLAR ==========

def xabar_yuborish(from_id, from_ism, from_role, from_org_id, to_ids, subject, xabar, org_id, rasm_id=None, doc_id=None):
    """Xabarni bir yoki ko'p odamlarga yuborish"""
    try:
        conn = connect(); cur = conn.cursor()
        vaqt = hozir().strftime("%Y-%m-%d %H:%M:%S")

        # Asosiy xabar yaratish
        cur.execute('''INSERT INTO xabarlar(from_user_id,from_user_name,from_role,from_org_id,
                      subject,xabar,rasm_id,dokument_id,yaratilgan,org_id)
                      VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id''',
                    (from_id, from_ism, from_role, from_org_id, subject, xabar, rasm_id, doc_id, vaqt, org_id))

        xabar_id = cur.fetchone()[0]

        # Qabul qiluvchilarni qo'shish
        if isinstance(to_ids, list):
            for to_id, to_ism, to_role, to_org in to_ids:
                cur.execute('''INSERT INTO xabar_qabul_qiluvchilar(xabar_id,qabul_qiluvchi_id,qabul_qiluvchi_nomi,qabul_qiluvchi_role)
                              VALUES(%s,%s,%s,%s)''',
                            (xabar_id, to_id, to_ism, to_role))
        else:
            to_id, to_ism, to_role, to_org = to_ids
            cur.execute('''INSERT INTO xabar_qabul_qiluvchilar(xabar_id,qabul_qiluvchi_id,qabul_qiluvchi_nomi,qabul_qiluvchi_role)
                          VALUES(%s,%s,%s,%s)''',
                        (xabar_id, to_id, to_ism, to_role))

        conn.commit(); cur.close(); conn.close()
        return xabar_id
    except Exception as e:
        if 'conn' in locals():
            conn.rollback(); cur.close(); conn.close()
        raise Exception(f"Xabar yuborish xatosi: {str(e)}")

def xabar_inbox_olish(user_id, org_id):
    """Foydalanuvchi uchun kirgan xabarlarni olish"""
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT q.id,x.id,x.from_user_name,x.from_role,x.subject,x.xabar,
                   x.yaratilgan,q.holat
                   FROM xabar_qabul_qiluvchilar q
                   JOIN xabarlar x ON q.xabar_id=x.id
                   WHERE q.qabul_qiluvchi_id=%s AND x.org_id=%s
                   ORDER BY x.yaratilgan DESC''', (user_id, org_id))
    r = cur.fetchall(); cur.close(); conn.close()
    return r

def xabar_history_olish(org_id):
    """Tashkilotning barcha xabar tarixini olish"""
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT id,from_user_name,from_role,subject,xabar,yaratilgan FROM xabarlar
                   WHERE org_id=%s ORDER BY yaratilgan DESC LIMIT 100''', (org_id,))
    r = cur.fetchall(); cur.close(); conn.close()
    return r

def xabar_o_qish(recipient_id, xabar_id):
    """Xabarni o'qilgan deb belgilash"""
    try:
        conn = connect(); cur = conn.cursor()
        vaqt = hozir().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute('''UPDATE xabar_qabul_qiluvchilar SET holat=%s, o_qilgan=%s
                      WHERE qabul_qiluvchi_id=%s AND xabar_id=%s''',
                    ('o\'qildi', vaqt, recipient_id, xabar_id))
        conn.commit(); cur.close(); conn.close()
        return True
    except Exception as e:
        if 'conn' in locals():
            conn.rollback(); cur.close(); conn.close()
        raise Exception(f"Xabar o'qish xatosi: {str(e)}")

def xabar_search(org_id, keyword):
    """Xabarlarni qidiruv"""
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT id,from_user_name,from_role,subject,xabar,yaratilgan
                   FROM xabarlar
                   WHERE org_id=%s AND (subject ILIKE %s OR xabar ILIKE %s)
                   ORDER BY yaratilgan DESC''',
                (org_id, f"%{keyword}%", f"%{keyword}%"))
    r = cur.fetchall(); cur.close(); conn.close()
    return r

def xabar_olish(xabar_id):
    """Bitta xabarni toliq ma'lumot bilan olish"""
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT id,from_user_name,from_role,subject,xabar,rasm_id,dokument_id,yaratilgan
                   FROM xabarlar WHERE id=%s''', (xabar_id,))
    r = cur.fetchone(); cur.close(); conn.close()
    return r

def xabar_qabul_qiluvchilar(xabar_id):
    """Xabarning barcha qabul qiluvchilarini olish"""
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT qabul_qiluvchi_nomi,qabul_qiluvchi_role,holat,o_qilgan
                   FROM xabar_qabul_qiluvchilar WHERE xabar_id=%s''', (xabar_id,))
    r = cur.fetchall(); cur.close(); conn.close()
    return r

# ========== KIRM (INCOME) ==========

def kirm_qoshish(xodim_id, komp_id, turi, summa, izoh, sana, vaqt, created_by_role, created_by_name):
    """Xodimga kirm qo'shish"""
    try:
        conn = connect(); cur = conn.cursor()
        yaratilgan = hozir().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute('''INSERT INTO kirm(xodim_id,komp_id,turi,summa,izoh,sana,vaqt,yaratilgan,UPDATE_BY_ROLE,UPDATE_BY_NAME)
                      VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id''',
                    (xodim_id, komp_id, turi, summa, izoh, sana, vaqt, yaratilgan, created_by_role, created_by_name))
        kirm_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return kirm_id
    except Exception as e:
        if 'conn' in locals():
            conn.rollback(); cur.close(); conn.close()
        raise Exception(f"Kirm qo'shish xatosi: {str(e)}")

def xodim_funksiya_olish(xodim_id):
    """Xodimning shaxsiy funksiya sozlamalarini olish"""
    conn = connect(); cur = conn.cursor()
    cur.execute('''SELECT xod_gps,xod_selfie,xod_audio,xod_matn,
                          xod_face_id,xod_hikvision,xod_live_gps,xod_wifi
                   FROM xodimlar WHERE id=%s''', (xodim_id,))
    r = cur.fetchone(); cur.close(); conn.close()
    return r if r else (None,)*8

def xodim_funksiya_ozgartir(xodim_id, ustun, qiymat):
    """Xodim funksiyasini yoqish/o'chirish/nullga qaytarish"""
    conn = connect(); cur = conn.cursor()
    cur.execute(f"UPDATE xodimlar SET {ustun}=%s WHERE id=%s", (qiymat, xodim_id))
    conn.commit(); cur.close(); conn.close()

def xodim_kpi_olish(xodim_id):
    """Xodimning KPI sozlamalarini olish, yo'q bo'lsa default qaytaradi"""
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT * FROM xodim_kpi WHERE xodim_id=%s", (xodim_id,))
    r = cur.fetchone(); cur.close(); conn.close()
    return r

def xodim_kpi_yaratish(xodim_id, komp_id):
    """KPI yozuvi yaratish (mavjud bo'lmasa)"""
    conn = connect(); cur = conn.cursor()
    cur.execute("SELECT id FROM xodim_kpi WHERE xodim_id=%s", (xodim_id,))
    if not cur.fetchone():
        cur.execute("INSERT INTO xodim_kpi(xodim_id,komp_id) VALUES(%s,%s)", (xodim_id, komp_id))
        conn.commit()
    cur.close(); conn.close()

def xodim_kpi_yangilash(xodim_id, maydon, qiymat):
    """KPI maydonini yangilash"""
    conn = connect(); cur = conn.cursor()
    cur.execute(f"UPDATE xodim_kpi SET {maydon}=%s WHERE xodim_id=%s", (qiymat, xodim_id))
    conn.commit(); cur.close(); conn.close()

def kirm_holat_yangilash(kirm_id, holat):
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE kirm SET holat=%s WHERE id=%s", (holat, kirm_id))
    conn.commit(); cur.close(); conn.close()

def chiqim_holat_yangilash(chiqim_id, holat):
    conn = connect(); cur = conn.cursor()
    cur.execute("UPDATE chiqim SET holat=%s WHERE id=%s", (holat, chiqim_id))
    conn.commit(); cur.close(); conn.close()

def kirm_olish(xodim_id, komp_id, sana_from=None, sana_to=None):
    """Xodimning kirm yozuvlarini olish"""
    conn = connect(); cur = conn.cursor()
    if sana_from and sana_to:
        cur.execute('''SELECT id,turi,summa,izoh,sana,vaqt,yaratilgan,UPDATE_BY_NAME
                      FROM kirm WHERE xodim_id=%s AND komp_id=%s AND sana BETWEEN %s AND %s
                      ORDER BY sana DESC''',
                    (xodim_id, komp_id, sana_from, sana_to))
    else:
        cur.execute('''SELECT id,turi,summa,izoh,sana,vaqt,yaratilgan,UPDATE_BY_NAME
                      FROM kirm WHERE xodim_id=%s AND komp_id=%s
                      ORDER BY sana DESC LIMIT 30''',
                    (xodim_id, komp_id))
    r = cur.fetchall(); cur.close(); conn.close()
    return r

def kirm_jami(xodim_id, komp_id, sana_from=None, sana_to=None):
    """Xodimning umumiy kirm summasini hisoblash (faqat tasdiqlangan)"""
    conn = connect(); cur = conn.cursor()
    if sana_from and sana_to:
        cur.execute('''SELECT SUM(summa) FROM kirm
                      WHERE xodim_id=%s AND komp_id=%s AND sana BETWEEN %s AND %s
                      AND holat='qabul_qilindi' ''',
                    (xodim_id, komp_id, sana_from, sana_to))
    else:
        cur.execute('''SELECT SUM(summa) FROM kirm
                      WHERE xodim_id=%s AND komp_id=%s AND holat='qabul_qilindi' ''',
                    (xodim_id, komp_id))
    r = cur.fetchone(); cur.close(); conn.close()
    return float(r[0]) if r and r[0] else 0.0

# ========== CHIQIM (EXPENSE/PENALTY) ==========

def chiqim_qoshish(xodim_id, komp_id, turi, summa, izoh, sana, vaqt, created_by_role, created_by_name):
    """Xodimga chiqim qo'shish"""
    try:
        conn = connect(); cur = conn.cursor()
        yaratilgan = hozir().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute('''INSERT INTO chiqim(xodim_id,komp_id,turi,summa,izoh,sana,vaqt,yaratilgan,UPDATE_BY_ROLE,UPDATE_BY_NAME)
                      VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id''',
                    (xodim_id, komp_id, turi, summa, izoh, sana, vaqt, yaratilgan, created_by_role, created_by_name))
        chiqim_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return chiqim_id
    except Exception as e:
        if 'conn' in locals():
            conn.rollback(); cur.close(); conn.close()
        raise Exception(f"Chiqim qo'shish xatosi: {str(e)}")

def chiqim_olish(xodim_id, komp_id, sana_from=None, sana_to=None):
    """Xodimning chiqim yozuvlarini olish"""
    conn = connect(); cur = conn.cursor()
    if sana_from and sana_to:
        cur.execute('''SELECT id,turi,summa,izoh,sana,vaqt,yaratilgan,UPDATE_BY_NAME
                      FROM chiqim WHERE xodim_id=%s AND komp_id=%s AND sana BETWEEN %s AND %s
                      ORDER BY sana DESC''',
                    (xodim_id, komp_id, sana_from, sana_to))
    else:
        cur.execute('''SELECT id,turi,summa,izoh,sana,vaqt,yaratilgan,UPDATE_BY_NAME
                      FROM chiqim WHERE xodim_id=%s AND komp_id=%s
                      ORDER BY sana DESC LIMIT 30''',
                    (xodim_id, komp_id))
    r = cur.fetchall(); cur.close(); conn.close()
    return r

def chiqim_jami(xodim_id, komp_id, sana_from=None, sana_to=None):
    """Xodimning umumiy chiqim summasini hisoblash (faqat tasdiqlangan)"""
    conn = connect(); cur = conn.cursor()
    if sana_from and sana_to:
        cur.execute('''SELECT SUM(summa) FROM chiqim
                      WHERE xodim_id=%s AND komp_id=%s AND sana BETWEEN %s AND %s
                      AND holat='qabul_qilindi' ''',
                    (xodim_id, komp_id, sana_from, sana_to))
    else:
        cur.execute('''SELECT SUM(summa) FROM chiqim
                      WHERE xodim_id=%s AND komp_id=%s AND holat='qabul_qilindi' ''',
                    (xodim_id, komp_id))
    r = cur.fetchone(); cur.close(); conn.close()
    return float(r[0]) if r and r[0] else 0.0
